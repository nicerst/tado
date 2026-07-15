#!/usr/bin/env python3
"""Export VAPT-BASELINE.md tables to a multi-sheet Excel checklist.

Usage: export_checklist.py <path-to-VAPT-BASELINE.md> <output.xlsx>
"""
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

SEVERITY_FILLS = {
    "CRITICAL": "FFC7CE",
    "HIGH": "FFD9B3",
    "MEDIUM": "FFF2CC",
    "LOW": "C6E0B4",
    "INFO": "DDEBF7",
}

SEVERITY_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]

STRIDE_FILLS = {
    "✓": "C6E0B4",
    "⚠": "FFF2CC",
    "✗": "FFC7CE",
}

STRIDE_NAMES = {
    "S": "Spoofing",
    "T": "Tampering",
    "R": "Repudiation",
    "I": "Information Disclosure",
    "D": "Denial of Service",
    "E": "Elevation of Privilege",
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def split_row(line):
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    cells = re.split(r"(?<!\\)\|", line)
    return [c.strip().replace("\\|", "|") for c in cells]


def parse_table(lines, start):
    """Parse a markdown table starting at `start` (the header row).

    Returns (header_cells, data_rows, next_index).
    """
    header = split_row(lines[start])
    rows = []
    i = start + 2  # skip header + separator row
    while i < len(lines) and lines[i].strip().startswith("|"):
        cells = split_row(lines[i])
        if len(cells) == len(header):
            rows.append(cells)
        i += 1
    return header, rows, i


def find_section_table(lines, heading):
    for i, line in enumerate(lines):
        if line.strip() == heading:
            for j in range(i + 1, len(lines)):
                if lines[j].strip().startswith("|"):
                    return parse_table(lines, j)
                if lines[j].strip().startswith("## "):
                    break
    return None, [], None


def find_section_text(lines, heading):
    out = []
    capture = False
    for line in lines:
        if line.strip() == heading:
            capture = True
            continue
        if capture:
            if line.strip().startswith("## "):
                break
            out.append(line.rstrip())
    return out


def severity_key(cell):
    upper = str(cell).upper()
    for key in SEVERITY_FILLS:
        if key in upper:
            return key
    return None


def severity_rank(cell):
    key = severity_key(cell)
    return SEVERITY_ORDER.index(key) if key in SEVERITY_ORDER else len(SEVERITY_ORDER)


def split_location(text):
    """Split a 'Location / Evidence' cell into (file, line(s)) from its first backtick token."""
    m = re.match(r"`([^`]+)`", text.strip())
    if not m:
        return "", ""
    token = m.group(1)
    fm = re.match(r"^(.*?):([\d,\-\s]+)$", token)
    if fm and re.search(r"\d", fm.group(2)):
        return fm.group(1), fm.group(2)
    return token, ""


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, ncols, rows, max_width=45):
    for col in range(1, ncols + 1):
        longest = max(
            [len(str(ws.cell(row=1, column=col).value or ""))]
            + [len(str(r[col - 1])) for r in rows if col - 1 < len(r)]
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(
            max(longest + 2, 10), max_width
        )


def build_findings_sheet(wb, header, rows):
    ws = wb.create_sheet("Findings Checklist")

    # Split the combined Location/Evidence column into File / Line(s) / Location Detail
    # so the dev can jump to code without parsing a long mixed cell.
    loc_idx = next((i for i, n in enumerate(header) if "location" in n.strip().lower()), None)
    if loc_idx is not None:
        new_header = header[:loc_idx] + ["File", "Line(s)", "Location Detail"] + header[loc_idx + 1:]
        new_rows = []
        for row in rows:
            cell = str(row[loc_idx]) if loc_idx < len(row) else ""
            file_part, line_part = split_location(cell)
            new_rows.append(row[:loc_idx] + [file_part, line_part, cell] + row[loc_idx + 1:])
        header, rows = new_header, new_rows

    extra_cols = ["Done", "Owner", "Due Date", "Notes"]
    out_header = ["#"] + header + extra_cols
    ws.append(out_header)

    for idx, row in enumerate(rows, start=1):
        ws.append([idx] + row + ["", "", "", ""])

    style_header(ws, len(out_header))

    sev_col = None
    for c, name in enumerate(header, start=2):  # +1 for "#" col, 1-indexed
        if name.strip().lower() == "severity":
            sev_col = c
            break

    if sev_col:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=sev_col)
            key = severity_key(str(cell.value or ""))
            if key:
                cell.fill = PatternFill("solid", fgColor=SEVERITY_FILLS[key])

    # Wrap long text + top-align so rows read cleanly without manual resizing.
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(out_header) + 1):
            ws.cell(row=r, column=c).alignment = WRAP_TOP

    done_col = len(out_header) - 3
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(row=2, column=done_col).coordinate}:{ws.cell(row=max(ws.max_row, 2), column=done_col).coordinate}")

    autosize(ws, len(out_header), rows)


def build_threat_model_sheet(wb, header, rows):
    if not rows:
        return
    ws = wb.create_sheet("Threat Model")
    header = [STRIDE_NAMES.get(h.strip(), h) for h in header]
    out_header = header + ["Notes"]
    ws.append(out_header)

    for row in rows:
        ws.append(row + [""])

    style_header(ws, len(out_header))

    for r in range(2, ws.max_row + 1):
        for c in range(2, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            text = str(cell.value or "")
            for symbol, fill in STRIDE_FILLS.items():
                if text.startswith(symbol):
                    cell.fill = PatternFill("solid", fgColor=fill)
                    break

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(out_header) + 1):
            ws.cell(row=r, column=c).alignment = WRAP_TOP

    autosize(ws, len(out_header), rows)


def build_scope_sheet(wb, scope_lines, gaps_lines, generated):
    ws = wb.create_sheet("Scope & Meta")
    ws.append(["Field", "Value"])
    style_header(ws, 2)

    if generated:
        ws.append(["Generated", generated])

    for line in scope_lines:
        m = re.match(r"-\s*([^:]+):\s*(.*)", line.strip())
        if m:
            ws.append([m.group(1).strip(), m.group(2).strip()])

    if gaps_lines:
        ws.append(["", ""])
        ws.append(["Tooling Gaps", ""])
        for line in gaps_lines:
            text = line.strip().lstrip("-").strip()
            if text:
                ws.append(["", text])

    for r in range(2, ws.max_row + 1):
        for c in (1, 2):
            ws.cell(row=r, column=c).alignment = WRAP_TOP

    autosize(ws, 2, [], max_width=80)


def build_summary_sheet(ws, project_name, generated, risk_score_line, header, rows):
    title = "VAPT Summary"
    if project_name:
        title += f" — {project_name}"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)

    if generated:
        ws.append([f"Generated: {generated}"])
    if risk_score_line:
        ws.append([f"Risk Score: {risk_score_line}"])

    ws.append([])

    # Severity counts table
    sev_idx = next((i for i, n in enumerate(header) if n.strip().lower() == "severity"), None)
    counts = {key: 0 for key in SEVERITY_ORDER}
    for row in rows:
        if sev_idx is not None and sev_idx < len(row):
            key = severity_key(row[sev_idx])
            if key in counts:
                counts[key] += 1

    ws.append(["Severity", "Count"])
    header_row = ws.max_row
    for col in (1, 2):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for key in SEVERITY_ORDER:
        ws.append([key, counts[key]])
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=SEVERITY_FILLS[key])

    ws.append([])
    ws.append(["Start here (highest severity first)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    finding_idx = next((i for i, n in enumerate(header) if n.strip().lower() == "finding"), None)
    for n, row in enumerate(rows[:5], start=1):
        sev = row[sev_idx] if sev_idx is not None and sev_idx < len(row) else ""
        finding_text = str(row[finding_idx]) if finding_idx is not None and finding_idx < len(row) else ""
        short = finding_text.split(". ")[0].split(".\n")[0]
        if len(short) > 110:
            short = short[:110] + "…"
        ws.append([f"#{n}", f"{sev}  {short}"])
        ws.cell(row=ws.max_row, column=2).alignment = WRAP_TOP

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100


def main():
    if len(sys.argv) != 3:
        print("usage: export_checklist.py <VAPT-BASELINE.md> <output.xlsx>")
        sys.exit(1)

    src, dest = sys.argv[1], sys.argv[2]
    with open(src, encoding="utf-8") as f:
        lines = f.readlines()

    generated = None
    project_name = None
    risk_score_line = None
    for line in lines:
        stripped = line.strip()
        m = re.match(r"_Generated:\s*([^|]+)", stripped)
        if m and generated is None:
            generated = m.group(1).strip()
        m = re.match(r"#\s+VAPT Baseline\s+—\s+(.+)", stripped)
        if m and project_name is None:
            project_name = m.group(1).strip()
        m = re.match(r"##\s+Risk Score:\s*(.+)", stripped)
        if m and risk_score_line is None:
            risk_score_line = m.group(1).strip()

    findings_header, findings_rows, _ = find_section_table(lines, "## Findings")
    if not findings_header:
        print("No '## Findings' table found in source file.")
        sys.exit(1)

    # Highest severity first so the dev sees what matters most at the top of the sheet.
    sev_idx = next((i for i, n in enumerate(findings_header) if n.strip().lower() == "severity"), None)
    if sev_idx is not None:
        findings_rows.sort(key=lambda r: severity_rank(r[sev_idx]) if sev_idx < len(r) else len(SEVERITY_ORDER))

    threat_header, threat_rows, _ = find_section_table(lines, "## Threat Model Snapshot")
    scope_lines = find_section_text(lines, "## Scope")
    gaps_lines = find_section_text(lines, "## Tooling Gaps")

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    build_findings_sheet(wb, findings_header, findings_rows)
    build_threat_model_sheet(wb, threat_header, threat_rows)
    build_scope_sheet(wb, scope_lines, gaps_lines, generated)
    build_summary_sheet(summary_ws, project_name, generated, risk_score_line, findings_header, findings_rows)

    wb.save(dest)
    print(f"Wrote {dest}")


if __name__ == "__main__":
    main()
